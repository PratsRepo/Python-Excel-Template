#
from xlutils.copy import copy
from xlrd import open_workbook
import xlwt

book=open_workbook("/Users/tusharbakaya/Desktop/test.xlsx")
book1=copy(book)
sheet=book.sheet_by_name('Sheet1')
sheet1=book1.get_sheet(0)
totalrows=sheet.nrows
print totalrows
for j in range(0,totalrows):
    for i in range(0,totalrows):
        row=sheet.cell_value(i,0)
        sheet1.write(i+(j*totalrows),0,row)
book1.save("/Users/tusharbakaya/Desktop/test1.xls")

# use a while loop instead of for loop and
from xlutils.copy import copy
from xlrd import open_workbook
import xlwt

book=open_workbook("/Users/tusharbakaya/Desktop/test.xlsx")
book1=copy(book)
sheet=book.sheet_by_name('Sheet1')
sheet1=book1.get_sheet(0)
totalrows=sheet.nrows
print totalrows
j,k = 0, 0
while k < totalrows
    for i in range(0,totalrows):
        row=sheet.cell_value(i,0)
        sheet1.write(i+j,0,row)
    j+=totalrows
    k += 1
book1.save("/Users/tusharbakaya/Desktop/test1.xls")

#Copy data from one Excel workbook to a new workbook with Python & OpenPyXL
# https://connysoderholm.com/copy-data-from-one-excel-workbook-to-a-new-workbook-with-python-openpyxl/
#!/usr/bin/python
# -*- coding: utf-8 -*-

# Create WB2 sheets WS1-WS10
for i in range(1, 11):
    WB2.create_sheet(f"WS{i}")

# delete first sheet
WB2.remove(WB2.worksheets[0])

# Define the copy ranges and sheets
copy_ranges = [100, 200, 50, 300, 350]
copy_to_sheets = ["WS1", "WS2", "WS3", "WS4", "WS4"]

# Copying the data
for s in range(i):
    offset += copy_ranges[s]

# Now it is time to fill our sheets with data! we traverse through our offset range with a for loop and set the values of the corresponding sheet. First we get the row with for j in range(offset, offset + copy_ranges[i]):. Next up are the cells in each row:
for row in WB1_WS1.iter_rows(min_row=j, max_row=j, min_col=1, max_col=WB1_WS1.max_column):

# We get the values for values_row with a list comprehension [cell.value for cell in row]. Finally, we append the row to the sheet with ws.append(values_row).
# Copy the row with the help of iter_rows, append the row
for j in range(offset, offset + copy_ranges[i]):
    # if j == 0:
    #    continue
    for row in WB1_WS1.iter_rows(min_row=j, max_row=j, min_col=1, max_col=WB1_WS1.max_column):
        values_row = [cell.value for cell in row]
    ws.append(values_row)

"""
Could you please suggest  how to copy the data from on work book to other book with specified rows
Source: Excel work book "WB1" having work sheet "WS1", This sheet  having 1000 rows of data
Destination: New work book 'WB2' and  work sheets WS1,WS2...WS10
Could you please suggest the code for following condition:
Copy the first 100 rows data and paste it WS1 sheet
Copy the next 200 rows data and paste it WS2 sheet
Copy the next 50 rows data and paste it WS3 sheet
Copy the next 300 rows data and paste it WS4 sheet
Copy the next 350 rows data and paste it WS4 sheet
"""

from openpyxl import Workbook, load_workbook

WB1 = load_workbook("Source.xlsx", data_only=True)
WB1_WS1 = WB1["WS1"]
WB2 = Workbook()

# Create WB2 sheets WS1-WS10
for i in range(1, 11):
    WB2.create_sheet(f"WS{i}")

# delete first sheet
WB2.remove(WB2.worksheets[0])

# Define the copy ranges and sheets
copy_ranges = [100, 200, 50, 300, 350]
copy_to_sheets = ["WS1", "WS2", "WS3", "WS4", "WS4"]

# Copy the values from the rows in WB1 to WB2.
for i in range( len(copy_ranges)):
    # Set the sheet to copy to
    ws = WB2[ copy_to_sheets[i] ]
    # Initialize row offset
    offset = 1
    # Set the row offset
    for s in range(i):
        offset += copy_ranges[s]

    # Copy the row with the help of iter_rows, append the row
    for j in range(offset,  offset + copy_ranges[i]):
        #if j == 0:
        #    continue
        for row in WB1_WS1.iter_rows(min_row=j, max_row=j, min_col=1, max_col=WB1_WS1.max_column):
            values_row = [cell.value for cell in row]
        ws.append(values_row)

# Save the workbook
WB2.save("WB2.xlsx")