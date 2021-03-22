# Iterating by rows
# import load_workbook
from openpyxl import load_workbook

# set file path
filepath = "/home/ubuntu/demo.xlsx"
# load demo.xlsx
wb = load_workbook(filepath)
# select demo.xlsx
sheet = wb.active
# get max row count
max_row = sheet.max_row
# get max column count
max_column = sheet.max_column
# iterate over all cells
# iterate over all rows
for i in range(1, max_row + 1):

    # iterate over all columns
    for j in range(1, max_column + 1):
        # get particular cell value
        cell_obj = sheet.cell(row=i, column=j)
        # print cell value
        print(cell_obj.value, end=' | ')
    # print new line
    print('\n')



# Add a sheet to the existing xlsx
# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="/home/ubuntu/demo.xlsx"
# load demo.xlsx
wb=load_workbook(filepath)
# create new sheet
wb.create_sheet('Sheet 2')
# save workbook
wb.save(filepath)

# Copy data from one sheet to another sheet
# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="/home/ubuntu/demo.xlsx"
# load demo.xlsx
wb=load_workbook(filepath)
# get Sheet
source=wb.get_sheet_by_name('Sheet')
# copy sheet
target=wb.copy_worksheet(source)
# save workbook
wb.save(filepath)

# Remove sheet from existing xlsx
# import load_workbook
from openpyxl import load_workbook
# set file path
filepath="/home/ubuntu/demo.xlsx"
# load demo.xlsx
wb=load_workbook(filepath)
# create new sheet
wb.remove(wb.get_sheet_by_name('Sheet 2'))
# save workbook
wb.save(filepath)


# Read in the spreadsheet data
ps = openpyxl.load_workbook(‘produceSales.xlsx’)
sheet = ps[‘Sheet1’]
sheet.max_row
#returns the total number of rows in the sheet
23758

# Next, we use a For loop to iterate over all the rows in the sheet.
for row in range(2, sheet.max_row + 1):
# each row in the spreadsheet represents information for a particular purchase.
produce = sheet[‘B’ + str(row)].value
cost_per_pound = sheet[‘C’ + str(row)].value
pounds_sold = sheet[‘D’ + str(row)].value
total_sales = sheet[‘E’ + str(row)].value
# the first column is B followed by C and so on.
# Each value in a cell is represented by a column letter and a row number. So #the first element in the sheet is B1, next column C1 and so on. This enables #to iterate over the entire cells.

# follow-- https://medium.com/analytics-vidhya/how-to-extract-information-from-your-excel-sheet-using-python-5f4f518aec49